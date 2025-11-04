import os

import openpyxl

from excel_framework.config.setting import DATA_PATH


def get_row_values(file_name, sheet_name, row_num):
    """
    获取指定行的结构化数据（字典格式，key=表头）
    :param file_name: Excel文件名（如test.xlsx）
    :param sheet_name: 工作表名（如add）
    :param row_num: 行号（1=表头，2=第一条用例）
    :return: 行数据字典
    """
    file_path = os.path.join(DATA_PATH, file_name)
    # file_path = DATA_PATH / file_name
    print("实际读取的文件路径：", file_path)
    print(f"读取：文件={file_name} | sheet={sheet_name} | 行号={row_num}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel文件不存在：{file_path}")
    if not file_path.endswith('.xlsx'):
        raise ValueError(f"文件格式错误：{file_path}")

    # 加载excel数据文件
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"sheet不存在：{sheet_name}")
        # 具体某一sheet
        sheet = workbook[sheet_name]

        # 校验行号有效性
        if row_num < 1 or row_num > sheet.max_row:
            raise IndexError(f"行号超出范围：{row_num}")

        # 读取表头
        headers = []
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            headers.append(header_value if header_value else f"col_{col}")

        # 读取目标数据
        row_data = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_num, column=col).value
            # 特殊处理：param列 （JSON字符串转字典，方便后续接口请求使用）
            if row_num > 1 and headers[col - 1] == 'param' and cell_value:
                try:
                    cell_value = eval(cell_value)
                except Exception:
                    print(f"警告：行{row_num}的param列不是有效JSON格式，保持原字符串")
            # 赋值到字典（key=表头，value=单元值）
            row_data[headers[col - 1]] = cell_value

        return row_data

    finally:
        workbook.close()


def get_test_data(file_name, sheet_name, skip_header=True):
    """
    获取工作表所有数据（结构化字典列表）
    :param file_name: Excel文件名
    :param sheet_name: 工作表名
    :param skip_header: 是否跳过表头（默认True，仅返回用例数据）
    :return: 数据列表（每个元素是一行数据字典）
    """
    file_path = os.path.join(DATA_PATH, file_name)
    print(f"批量读取：文件={file_name} | sheet={sheet_name}（跳过表头：{skip_header}）")
    # file_path = DATA_PATH / file_name
    # 加载excel数据文件,仅加载一次文件，提升性能
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    # 具体某一sheet
    sheet = workbook[sheet_name]
    data = []
    # 行数
    rows = sheet.max_row

    for j in range(1, rows + 1):
        data.append(get_row_values(file_name, sheet_name, j))
    return data


if __name__ == '__main__':
    print(get_row_values('test.xlsx', 'add', 1))
    try:
        # 正确写法：调用函数并打印返回值
        row_data = get_row_values('test.xlsx', 'add', 1)
        print("第1行数据：", row_data)  # 打印函数返回的行数据
    except Exception as e:
        print("执行错误：", e)  # 打印异常信息（而非路径）